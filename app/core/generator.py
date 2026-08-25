from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.units import pixels_to_EMU


# ============================================================
# TOKENS
# ============================================================

def _replace_tokens(value: Any, variables: dict) -> Any:
    if not isinstance(value, str):
        return value

    out = value

    for key, val in variables.items():
        token = "{{" + str(key) + "}}"
        replacement = "" if val is None else str(val)
        out = out.replace(token, replacement)

    return out


# ============================================================
# ESTILO
# ============================================================

def _copy_row_style(
    ws,
    source_row: int,
    target_row: int,
    max_col: int,
):
    for col in range(1, max_col + 1):

        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if isinstance(src, MergedCell):
            continue

        if isinstance(dst, MergedCell):
            continue

        if src.has_style:
            dst._style = copy(src._style)

        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format

    source_height = ws.row_dimensions[source_row].height

    if source_height is not None:
        ws.row_dimensions[target_row].height = source_height


# ============================================================
# CORES POR TIPO DE DESLOCAMENTO
# ============================================================
#
# ENTRADA
# -> VERDE
#
# FECHAMENTO
# -> AZUL
#
# ESPECIAL
# -> AMARELO
#
# DESLOCAMENTO_RECARGA
# -> LARANJA
#
# Para deslocamento de recarga, toda a faixa H:U
# da linha deve ficar laranja.
#
# Se o tipo for null/vazio, nenhuma alteração de cor é feita.
# O estilo original do template é preservado.
# ============================================================

FILL_ENTRADA = PatternFill(
    fill_type="solid",
    fgColor="C4D79B",
)

FILL_FECHAMENTO = PatternFill(
    fill_type="solid",
    fgColor="9FC5E8",
)

FILL_ESPECIAL = PatternFill(
    fill_type="solid",
    fgColor="FFFF66",
)

FILL_RECARGA = PatternFill(
    fill_type="solid",
    fgColor="FABF8F",
)


def _get_displacement_fill(
    displacement_type: Any,
):
    """
    Retorna a cor correspondente ao tipo de deslocamento.

    ENTRADA                -> verde
    FECHAMENTO             -> azul
    ESPECIAL               -> amarelo
    DESLOCAMENTO_RECARGA   -> laranja

    Para None, vazio ou qualquer outro valor,
    não altera a formatação existente.
    """

    if displacement_type is None:
        return None

    displacement_type = str(
        displacement_type
    ).strip().upper()

    if not displacement_type:
        return None

    if displacement_type == "ENTRADA":
        return FILL_ENTRADA

    if displacement_type == "FECHAMENTO":
        return FILL_FECHAMENTO

    if displacement_type == "ESPECIAL":
        return FILL_ESPECIAL

    if displacement_type == "DESLOCAMENTO_RECARGA":
        return FILL_RECARGA

    return None


def _apply_displacement_color(
    ws,
    row_num: int,
    start_col: int,
    end_col: int,
    displacement_type: Any,
):
    """
    Aplica a cor do tipo de deslocamento somente
    nas células preenchidas do quadrante.

    As células vazias mantêm o estilo original
    copiado do template.
    """

    fill = _get_displacement_fill(
        displacement_type
    )

    if fill is None:
        return

    for col_num in range(
        start_col,
        end_col + 1,
    ):

        cell = ws.cell(
            row=row_num,
            column=col_num,
        )

        if isinstance(
            cell,
            MergedCell,
        ):
            continue

        if (
            cell.value is None
            or cell.value == ""
        ):
            continue

        cell.fill = copy(fill)


def _apply_recharge_color(
    ws,
    row_num: int,
    displacement_type: Any,
):
    """
    Quando o Quadrante 3 representa um
    DESLOCAMENTO_RECARGA, aplica a cor laranja
    em toda a faixa H:U da mesma linha.

    A regra é aplicada por último para que tenha
    prioridade sobre ENTRADA, FECHAMENTO e ESPECIAL.

    Diferentemente da regra padrão de deslocamento,
    aqui TODAS as células entre H e U são coloridas,
    inclusive células vazias.
    """

    if displacement_type is None:
        return

    displacement_type = str(
        displacement_type
    ).strip().upper()

    if (
        displacement_type
        != "DESLOCAMENTO_RECARGA"
    ):
        return

    for col_num in range(
        8,
        22,
    ):

        cell = ws.cell(
            row=row_num,
            column=col_num,
        )

        if isinstance(
            cell,
            MergedCell,
        ):
            continue

        cell.fill = copy(
            FILL_RECARGA
        )


# ============================================================
# MARCADORES
# ============================================================

def _find_marker(ws, marker: str):

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell, MergedCell):
                continue

            if cell.value == marker:
                return cell.row, cell.column

    return None


# ============================================================
# VARIÁVEIS
# ============================================================

def _replace_variables(ws, variables: dict):

    if not variables:
        return

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell, MergedCell):
                continue

            if isinstance(cell.value, str):
                cell.value = _replace_tokens(
                    cell.value,
                    variables,
                )


# ============================================================
# NOME DA ABA
# ============================================================

def _replace_sheet_title(ws, variables: dict):

    new_title = _replace_tokens(
        ws.title,
        variables,
    )

    if new_title == ws.title:
        return

    invalid_chars = [
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]",
    ]

    for char in invalid_chars:
        new_title = new_title.replace(char, "-")

    new_title = new_title[:31].strip()

    if not new_title:
        new_title = "Relatorio"

    ws.title = new_title


# ============================================================
# MERGES
# ============================================================

def _capture_and_remove_merges_below(
    ws,
    insertion_row: int,
):
    merges_to_move = []

    for merged_range in list(ws.merged_cells.ranges):

        if merged_range.min_row >= insertion_row:

            merges_to_move.append(
                (
                    merged_range.min_col,
                    merged_range.min_row,
                    merged_range.max_col,
                    merged_range.max_row,
                )
            )

            ws.unmerge_cells(
                str(merged_range)
            )

    return merges_to_move


def _restore_shifted_merges(
    ws,
    merges,
    row_offset: int,
):
    if row_offset <= 0:
        return

    for (
        min_col,
        min_row,
        max_col,
        max_row,
    ) in merges:

        new_min_row = min_row + row_offset
        new_max_row = max_row + row_offset

        start_cell = (
            f"{get_column_letter(min_col)}"
            f"{new_min_row}"
        )

        end_cell = (
            f"{get_column_letter(max_col)}"
            f"{new_max_row}"
        )

        ws.merge_cells(
            f"{start_cell}:{end_cell}"
        )


# ============================================================
# GARANTIA: REGISTROS API SEM MERGES
# ============================================================

def _remove_merges_from_data_area(
    ws,
    start_row: int,
    record_count: int,
    max_data_col: int,
):
    if record_count <= 0:
        return

    end_row = (
        start_row
        + record_count
        - 1
    )

    for merged_range in list(
        ws.merged_cells.ranges
    ):

        row_overlap = not (
            merged_range.max_row < start_row
            or merged_range.min_row > end_row
        )

        col_overlap = (
            merged_range.min_col <= max_data_col
        )

        if row_overlap and col_overlap:
            ws.unmerge_cells(
                str(merged_range)
            )


# ============================================================
# IMAGENS EXISTENTES
# ============================================================

def _shift_images_below(
    ws,
    insertion_row: int,
    row_offset: int,
):
    if row_offset <= 0:
        return

    for image in ws._images:

        anchor = image.anchor

        if not hasattr(anchor, "_from"):
            continue

        image_row = (
            anchor._from.row + 1
        )

        if image_row >= insertion_row:

            anchor._from.row += row_offset

            if hasattr(anchor, "_to"):
                anchor._to.row += row_offset


# ============================================================
# NOVAS IMAGENS CONFIGURADAS
# ============================================================

def _column_width_to_pixels(width: float) -> int:
    """
    Aproxima a largura de uma coluna do Excel em pixels.
    """
    if width is None:
        width = 8.43

    if width < 1:
        return int(width * 12)

    return int(width * 7 + 5)


def _add_configured_images(
    ws,
    images_config: dict,
    tenant_dir: Path,
):
    """
    Insere imagens definidas no config.yaml.

    O path é relativo à raiz do tenant.

    Exemplo:

    images:
      logo:
        path: "assets/logo.png"
        end_column: "U"
        row: 1
        width: 180
        height: 55
        horizontal_align: "right"
    """

    if not images_config:
        return

    for image_name, image_cfg in images_config.items():

        image_path = (
            tenant_dir
            / image_cfg["path"]
        ).resolve()

        if not image_path.exists():
            raise FileNotFoundError(
                f"Imagem '{image_name}' não encontrada: "
                f"{image_path}"
            )

        img = XLImage(
            str(image_path)
        )

        width = int(
            image_cfg.get(
                "width",
                img.width,
            )
        )

        height = int(
            image_cfg.get(
                "height",
                img.height,
            )
        )

        img.width = width
        img.height = height

        row = int(
            image_cfg.get(
                "row",
                1,
            )
        )

        end_column_letter = str(
            image_cfg.get(
                "end_column",
                "A",
            )
        ).upper()

        horizontal_align = str(
            image_cfg.get(
                "horizontal_align",
                "left",
            )
        ).lower()

        end_col = column_index_from_string(
            end_column_letter
        )

        # ----------------------------------------------------
        # ALINHAMENTO À DIREITA
        # ----------------------------------------------------

        if horizontal_align == "right":

            # Calcula quantos pixels existem da coluna A
            # até o final da coluna configurada.
            total_pixels = 0

            for col_index in range(
                1,
                end_col + 1,
            ):

                letter = get_column_letter(
                    col_index
                )

                column_width = (
                    ws.column_dimensions[
                        letter
                    ].width
                )

                total_pixels += (
                    _column_width_to_pixels(
                        column_width
                    )
                )

            # A imagem deve terminar exatamente
            # no final da coluna configurada.
            start_pixels = (
                total_pixels
                - width
            )

            if start_pixels < 0:
                start_pixels = 0

            # Descobre em qual coluna a imagem começa.
            accumulated_pixels = 0
            start_col = 1
            offset_pixels = 0

            for col_index in range(
                1,
                end_col + 1,
            ):

                letter = get_column_letter(
                    col_index
                )

                column_pixels = (
                    _column_width_to_pixels(
                        ws.column_dimensions[
                            letter
                        ].width
                    )
                )

                if (
                    accumulated_pixels
                    + column_pixels
                    > start_pixels
                ):
                    start_col = col_index
                    offset_pixels = (
                        start_pixels
                        - accumulated_pixels
                    )
                    break

                accumulated_pixels += (
                    column_pixels
                )

            marker = AnchorMarker(
                col=start_col - 1,
                colOff=pixels_to_EMU(
                    offset_pixels
                ),
                row=row - 1,
                rowOff=0,
            )

            size = XDRPositiveSize2D(
                cx=pixels_to_EMU(width),
                cy=pixels_to_EMU(height),
            )

            img.anchor = OneCellAnchor(
                _from=marker,
                ext=size,
            )

            ws.add_image(img)

        # ----------------------------------------------------
        # ALINHAMENTO SIMPLES
        # ----------------------------------------------------

        else:

            anchor = (
                f"{end_column_letter}"
                f"{row}"
            )

            ws.add_image(
                img,
                anchor,
            )


# ============================================================
# ÁREA DE IMPRESSÃO
# ============================================================

def _update_print_area(
    ws,
    inserted_rows: int,
    insertion_start_row: int,
):
    if inserted_rows <= 0:
        return

    print_area = ws.print_area

    if not print_area:
        return

    try:
        ranges = list(print_area.ranges)
    except Exception:
        return

    new_ranges = []

    for cell_range in ranges:

        min_col = cell_range.min_col
        min_row = cell_range.min_row
        max_col = cell_range.max_col
        max_row = cell_range.max_row

        if max_row >= insertion_start_row:
            max_row += inserted_rows

        start_cell = (
            f"{get_column_letter(min_col)}"
            f"{min_row}"
        )

        end_cell = (
            f"{get_column_letter(max_col)}"
            f"{max_row}"
        )

        new_ranges.append(
            f"{start_cell}:{end_cell}"
        )

    if new_ranges:
        ws.print_area = ",".join(
            new_ranges
        )


# ============================================================
# GERADOR
# ============================================================

def generate(
    template_path: Path,
    config: dict,
    payload: dict,
) -> BytesIO:

    template_path = Path(
        template_path
    )

    if not template_path.exists():
        raise FileNotFoundError(
            f"Template não encontrado: "
            f"{template_path}"
        )

    # Diretório do relatório:
    #
    # /app/tenants/sjc/reports/OSO
    #
    report_dir = (
        template_path
        .parent
        .parent
    )

    # Diretório do tenant:
    #
    # /app/tenants/sjc
    #
    tenant_dir = (
        report_dir
        .parent
        .parent
    )

    # ========================================================
    # ABRE WORKBOOK
    # ========================================================

    wb = load_workbook(
        template_path
    )

    # ========================================================
    # SELECIONA ABA
    # ========================================================

    configured_sheet = config.get(
        "sheet"
    )

    if configured_sheet:

        if configured_sheet not in wb.sheetnames:
            raise ValueError(
                f"A planilha '{configured_sheet}' "
                f"não existe no template. "
                f"Disponíveis: {wb.sheetnames}"
            )

        ws = wb[configured_sheet]

    else:
        ws = wb[
            wb.sheetnames[0]
        ]

    # ========================================================
    # REMOVE IMAGENS ORIGINAIS DO TEMPLATE
    # ========================================================
    #
    # Como agora as imagens são controladas pelo config.yaml,
    # removemos as imagens embutidas para evitar duplicidade.
    # ========================================================

    if config.get("images"):
        ws._images = []

    # ========================================================
    # VARIÁVEIS
    # ========================================================

    variables = (
        payload.get(
            "variables",
            {}
        )
        or {}
    )

    # ========================================================
    # SEÇÕES
    # ========================================================

    sections = (
        config.get(
            "sections",
            {}
        )
        or {}
    )

    found_sections = []

    for (
        section_name,
        section_cfg,
    ) in sections.items():

        marker = section_cfg.get(
            "marker",
            "{{"
            + section_name.upper()
            + "}}",
        )

        position = _find_marker(
            ws,
            marker,
        )

        if position is None:
            raise ValueError(
                f"Marcador '{marker}' "
                f"da seção '{section_name}' "
                f"não encontrado."
            )

        found_sections.append(
            (
                position[0],
                position[1],
                section_name,
                section_cfg,
            )
        )

    # ========================================================
    # PROCESSA SEÇÕES
    # ========================================================

    for (
        marker_row,
        marker_col,
        section_name,
        section_cfg,
    ) in sorted(
        found_sections,
        key=lambda item: item[0],
        reverse=True,
    ):

        records = (
            payload
            .get("sections", {})
            .get(section_name, [])
            or []
        )

        template_rows = (
            section_cfg.get(
                "template_rows"
            )
            or []
        )

        if not template_rows:

            old_template_row = (
                section_cfg.get(
                    "template_row"
                )
            )

            if old_template_row:
                template_rows = [
                    int(old_template_row)
                ]

        if not template_rows:
            template_rows = [
                marker_row
            ]

        template_rows = [
            int(row)
            for row in template_rows
        ]

        start_row = int(
            section_cfg.get(
                "start_row",
                template_rows[0],
            )
        )

        columns = (
            section_cfg.get(
                "columns",
                {}
            )
            or {}
        )

        marker_cell = ws.cell(
            marker_row,
            marker_col,
        )

        if not isinstance(
            marker_cell,
            MergedCell,
        ):
            marker_cell.value = None

        model_count = len(
            template_rows
        )

        record_count = len(
            records
        )

        if record_count == 0:

            for row_num in template_rows:

                for col in columns.values():

                    cell = ws.cell(
                        row_num,
                        int(col),
                    )

                    if not isinstance(
                        cell,
                        MergedCell,
                    ):
                        cell.value = None

            continue

        additional_rows = max(
            0,
            record_count - model_count,
        )

        insertion_row = (
            start_row
            + model_count
        )

        footer_merges = []

        if additional_rows > 0:

            footer_merges = (
                _capture_and_remove_merges_below(
                    ws,
                    insertion_row,
                )
            )

            ws.insert_rows(
                insertion_row,
                amount=additional_rows,
            )

            _restore_shifted_merges(
                ws,
                footer_merges,
                additional_rows,
            )

            _shift_images_below(
                ws,
                insertion_row,
                additional_rows,
            )

            _update_print_area(
                ws,
                additional_rows,
                insertion_row,
            )

        max_data_col = max(
            [
                int(col)
                for col
                in columns.values()
            ],
            default=1,
        )

        _remove_merges_from_data_area(
            ws,
            start_row,
            record_count,
            max_data_col,
        )

        # Zebrado
        for index in range(
            record_count
        ):

            target_row = (
                start_row
                + index
            )

            template_index = (
                index
                % model_count
            )

            source_row = (
                template_rows[
                    template_index
                ]
            )

            if target_row not in template_rows:

                _copy_row_style(
                    ws,
                    source_row,
                    target_row,
                    max_data_col,
                )

        # Dados
        for (
            index,
            record,
        ) in enumerate(records):

            row_num = (
                start_row
                + index
            )

            for (
                field,
                col,
            ) in columns.items():

                col_num = int(col)

                target_cell = ws.cell(
                    row_num,
                    col_num,
                )

                if isinstance(
                    target_cell,
                    MergedCell,
                ):
                    raise ValueError(
                        f"Célula "
                        f"{target_cell.coordinate} "
                        f"está mesclada dentro da "
                        f"área REGISTROS_API."
                    )

                target_cell.value = (
                    record.get(field)
                )

            # ====================================================
            # CORES DOS TIPOS DE DESLOCAMENTO
            #
            # Os campos abaixo são utilizados apenas como
            # metadados de formatação:
            #
            # 1_TIPO_DESLOCAMENTO
            # 2_TIPO_DESLOCAMENTO
            # 3_TIPO_DESLOCAMENTO
            # 4_TIPO_DESLOCAMENTO
            #
            # Eles NÃO precisam existir no config.yaml.
            # Eles NÃO são escritos em nenhuma célula.
            #
            # Quadrante 1 = A:G
            # Quadrante 2 = H:N
            # Quadrante 3 = O:Q
            # Quadrante 4 = R:U
            # ====================================================

            _apply_displacement_color(
                ws,
                row_num,
                1,
                7,
                record.get(
                    "1_TIPO_DESLOCAMENTO"
                ),
            )

            _apply_displacement_color(
                ws,
                row_num,
                8,
                14,
                record.get(
                    "2_TIPO_DESLOCAMENTO"
                ),
            )

            _apply_displacement_color(
                ws,
                row_num,
                15,
                17,
                record.get(
                    "3_TIPO_DESLOCAMENTO"
                ),
            )

            _apply_displacement_color(
                ws,
                row_num,
                18,
                21,
                record.get(
                    "4_TIPO_DESLOCAMENTO"
                ),
            )

            # ====================================================
            # COR DO DESLOCAMENTO DE RECARGA
            #
            # Quando o Q3 estiver identificado como
            # DESLOCAMENTO_RECARGA, toda a faixa H:U
            # deve ficar LARANJA.
            #
            # Esta regra é aplicada POR ÚLTIMO para ter
            # prioridade sobre:
            #
            # ENTRADA
            # FECHAMENTO
            # ESPECIAL
            #
            # ====================================================

            _apply_recharge_color(
                ws,
                row_num,
                record.get(
                    "3_TIPO_DESLOCAMENTO"
                ),
            )

    # ========================================================
    # SUBSTITUI TOKENS
    # ========================================================

    _replace_variables(
        ws,
        variables,
    )

    # ========================================================
    # NOME DA ABA
    # ========================================================

    _replace_sheet_title(
        ws,
        variables,
    )

    # ========================================================
    # ADICIONA IMAGENS CONFIGURADAS
    # ========================================================

    _add_configured_images(
        ws,
        config.get(
            "images",
            {}
        ),
        tenant_dir,
    )

    # ========================================================
    # SALVA
    # ========================================================

    out = BytesIO()

    wb.save(out)

    out.seek(0)

    return out
